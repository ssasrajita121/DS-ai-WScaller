"""
SnapSkill AI Caller - Backend Logic
Handles Vapi API integration with language-specific prompts
FIXED: Language codes for Deepgram compatibility
"""

import os
import re
import requests
from datetime import datetime
from dotenv import load_dotenv
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill

load_dotenv()

# ==========================================
# LANGUAGE-SPECIFIC CONFIGURATIONS
# ==========================================
LANGUAGE_CONFIG = {
   "Telugu (తెలుగు)": {
    "voice_provider": "azure",
    "voice_id": "te-IN-ShrutiNeural",
    "language_code": "multi",  # ← USE MULTI
    "voice_name": "Azure Telugu Voice - Shruti",
    "use_custom_api_key": False,
    "prompt": """నువ్వు SnapSkill నుండి feedback సేకరించే స్నేహపూర్వక సహాయకుడివి, తెలుగులో మాట్లాడతావు.

ముఖ్యమైన సూచనలు:
- మొత్తం సంభాషణ తెలుగులోనే చేయి
- సహజమైన తెలుగు మాట్లాడే భాషలో మాట్లాడు (ఇంగ్లీష్ నుండి అనువాదం లాగా కాదు)
- తెలుగు యొక్క సరైన ఇంటోనేషన్ మరియు లయ ఉపయోగించు
- ఇంగ్లీష్ పదాలు చాలా అవసరమైనప్పుడు మాత్రమే వాడు (Data Science వంటి technical terms)

సంభాషణ క్రమం:
1. శుభాకాంక్షలు: "నమస్కారం! నేను SnapSkill నుండి కాల్ చేస్తున్నాను."
2. ఉద్దేశ్యం: "మీరు మా Data Science ఫ్రీ కోర్స్ చేశారు కదా? మీ feedback తీసుకోవడానికి కాల్ చేశాను."
3. ప్రధాన ప్రశ్న: "కోర్స్ ఎలా ఉంది? మీకు నచ్చిందా?"
4. తదుపరి ప్రశ్నలు (సమాధానం ఆధారంగా అడుగు):
   - "కోర్స్‌లో ఏమి బాగా నచ్చింది?"
   - "ఏదైనా improve చేయాల్సిన అంశాలు ఉన్నాయా?"
   - "instructor teaching ఎలా ఉంది?"
   - "కోర్స్ materials అర్థం అయ్యాయా?"
   - "మీ స్నేహితులకు recommend చేస్తారా?"
5. ముగింపు: "మీ విలువైన feedback కోసం చాలా ధన్యవాదాలు! మా తదుపరి కోర్సులు గురించి మీకు తెలియజేస్తాం."

నియమాలు:
- తెలుగులో పూర్తిగా సహజంగా మాట్లాడు
- కాల్ 3 నిమిషాలలోపు ముగించు
- వెచ్చదనం మరియు అభినందనతో మాట్లాడు
- ప్రతికూల feedback ను సానుకూలంగా స్వీకరించు - కృతజ్ఞతలు చెప్పు
- గరిష్టంగా 2-3 తదుపరి ప్రశ్నలు అడుగు
- వ్యక్తి బిజీగా ఉంటే, తర్వాత కాల్ చేయమని చెప్పు
- వారి సూచనలను జాగ్రత్తగా వినండి

ఇప్పుడు సంభాషణ ప్రారంభించు. గుర్తుంచుకో - తెలుగులో ఆలోచించు మరియు తెలుగులో మాట్లాడు, ఇంగ్లీష్ నుండి translate చేయకు।"""
},
    
    "English": {
        "voice_provider": "11labs",
        "voice_id": "OUBnvvuqEKdDWtapoJFn",  # DS
        "language_code": "en-IN",
        "voice_name": "Bella - Indian English",
        # Voice parameters for professional Indian English
        "stability": 0.7,              # Higher = more consistent
        "similarity_boost": 0.75,      # Standard similarity
        "style": 0.3,                  # Lower style for professional tone
        "use_speaker_boost": True,
          # Natural English greeting 
        "prompt": """You are a professional feedback collection assistant calling from SnapSkill.
IMPORTANT INSTRUCTIONS:
- Your first message has already been said (the greeting)
- Wait for the customer's response before continuing
- Do NOT repeat the greeting

SCRIPT TO FOLLOW:
After the customer responds to your greeting, ask permission:
"You recently completed our Data Science free course. 
 May I take 2 minutes for your feedback?"
 Main Question: "How was your experience with the free project workshop? Did you find it helpful?"
Follow-up questions (based on response):
   - "What did you like most about the workshop?"
   - "Is there anything we can improve?"
   - "How would you rate the instructor's teaching?"
   - "Are you intrested in the detailed Data Science Course?"
Closing: "Thank you so much for your valuable feedback! We'll keep you updated about our upcoming courses."

RULES:
- Speak clearly in Indian English accent
- Keep call under 3 minutes
- Be professional yet friendly
- Accept negative feedback gracefully - thank them for honesty
- Ask 2-3 follow-up questions maximum
- If person is busy, offer to call back
- Listen actively to their suggestions

Start the conversation now."""
    },
    
    "Hindi (हिंदी)": {
        "voice_provider": "11labs",
        "voice_id": "Ms9OTvWb99V6DwRHZn6q",  # Your excellent voice! Update this
        "language_code": "hi",
        "voice_name": "ElevenLabs Hindi Voice",
        # Voice parameters for natural Indian accent
        "stability": 0.4,              # Lower = more expressive, natural
        "similarity_boost": 0.85,      # High = closer to source voice
        "style": 0.5,                  # Moderate style for natural flow
        "use_speaker_boost": True,     # Better speaker clarity
        "voice_language": "hi",        # CRITICAL: Force Hindi language model
        "prompt": """तुम SnapSkill की एक दोस्ताना feedback collection सहायक हो जो हिंदी में बात करती है।

महत्वपूर्ण निर्देश:
- पूरी बातचीत केवल हिंदी में करो
- प्राकृतिक हिंदी बोलचाल की भाषा में बात करो (अंग्रेजी से अनुवाद जैसा मत बोलो)
- हिंदी की उचित इंटोनेशन और लय का उपयोग करो
- अंग्रेजी शब्द तभी बोलो जब बहुत ज़रूरी हो (Data Science जैसे technical terms)

बातचीत का क्रम:
1. अभिवादन: "नमस्ते! मैं SnapSkill से बात कर रही हूं।"
2. उद्देश्य: "आपने हमारा Data Science का फ्री कोर्स किया था। आपका feedback लेने के लिए कॉल कर रही हूं।"
3. मुख्य सवाल: "कोर्स कैसा लगा? क्या आपको helpful लगा?"
4. अनुवर्ती सवाल (जवाब के आधार पर पूछो):
   - "कोर्स में सबसे अच्छा क्या लगा?"
   - "क्या कुछ improve करने की ज़रूरत है?"
   - "instructor की teaching कैसी थी?"
   - "कोर्स materials समझने में आसान थे?"
   - "क्या आप अपने दोस्तों को recommend करेंगे?"
5. समापन: "आपके valuable feedback के लिए बहुत-बहुत धन्यवाद! हम आपको आने वाले courses के बारे में बताते रहेंगे।"

नियम:
- हिंदी में पूरी तरह स्वाभाविक रूप से बोलो
- कॉल 3 मिनट से कम रखो
- गर्मजोशी और सराहना के साथ बात करो
- नकारात्मक feedback को सकारात्मक तरीके से स्वीकार करो - धन्यवाद दो
- ज़्यादा से ज़्यादा 2-3 अनुवर्ती सवाल पूछो
- अगर व्यक्ति व्यस्त है तो बाद में कॉल करने की पेशकश करो
- उनके सुझाव ध्यान से सुनो

अब बातचीत शुरू करो। याद रखो - हिंदी में सोचो और हिंदी में बोलो, अंग्रेजी से translate मत करो।""
- Accept negative feedback positively - thank them
- Ask 2-3 follow-up questions maximum
- If person is busy, offer to call back
- Note their suggestions

Start the conversation now."""
    }
}

# ==========================================
# VAPI API CONFIGURATION
# ==========================================
VAPI_API_KEY = os.getenv('VAPI_API_KEY')
VAPI_PHONE_NUMBER_ID = os.getenv('VAPI_PHONE_NUMBER_ID')
VAPI_BASE_URL = "https://api.vapi.ai"

# ==========================================
# PHONE NUMBER VALIDATION
# ==========================================
def validate_phone_number(phone):
    """
    Validate phone number format
    Returns: (is_valid, error_message)
    """
    if not phone:
        return False, "Phone number is required"
    
    # Remove spaces and special characters
    cleaned = re.sub(r'[^\d+]', '', phone)
    
    # Check if it starts with +91 and has 10 digits after
    if not cleaned.startswith('+91'):
        return False, "Phone number must start with +91"
    
    if len(cleaned) != 13:  # +91 + 10 digits
        return False, "Phone number must be +91 followed by 10 digits"
    
    # Check if the 10 digits are valid (start with 6-9)
    digits = cleaned[3:]
    if not digits[0] in '6789':
        return False, "Invalid Indian mobile number"
    
    return True, ""

# ==========================================
# CREATE VAPI ASSISTANT
# ==========================================
def create_assistant(voice_provider, voice_id, language_code, prompt, language_name, voice_params):
    """
    Create Vapi assistant with language-specific configuration
    """
    url = f"{VAPI_BASE_URL}/assistant"
    
    headers = {
        "Authorization": f"Bearer {VAPI_API_KEY}",
        "Content-Type": "application/json"
    }
    
    # Build voice config
    voice_config = {
        "provider": voice_provider,
        "voiceId": voice_id,
    }
    
    # Add ElevenLabs-specific parameters if using 11labs
    if voice_provider == "11labs":
        voice_config["stability"] = voice_params.get('stability', 0.5)
        voice_config["similarityBoost"] = voice_params.get('similarity_boost', 0.75)
        
        # CRITICAL: Add model parameter for language control
        if voice_params.get('voice_language'):
            voice_config["model"] = "eleven_multilingual_v2"  # Use multilingual model
            # Note: Vapi may use language_code from transcriber for voice language
        
        if 'style' in voice_params:
            voice_config["style"] = voice_params['style']
        if voice_params.get('use_speaker_boost'):
            voice_config["useSpeakerBoost"] = True
    
    payload = {
        "name": f"SnapSkill {language_name.split()[0]}",  # Max 40 chars
        "model": {
            "provider": "openai",
            "model": "gpt-4o",
            "messages": [
                {
                    "role": "system",
                    "content": prompt
                }
            ],
            "temperature": 0.7,
            "maxTokens": 500
        },
        "voice": voice_config,
        "firstMessage": "Hello! I'm calling from SnapSkill.!",
        "transcriber": {
            "provider": "deepgram",
            "model": "nova-2",  # Latest Deepgram model
            "language": language_code  # CRITICAL: Must be supported by nova-2
        },
        "recordingEnabled": True,
        "endCallMessage": "Thank you for your time. Goodbye!",
        "endCallPhrases": ["goodbye", "bye", "thank you bye", "not interested"]
    }
    
    print(f"\n🔧 Creating assistant for {language_name}...")
    print(f"   Provider: {voice_provider}")
    print(f"   Language code: {language_code}")
    print(f"   Voice ID: {voice_id[:20]}...")
    if voice_provider == "11labs":
        print(f"   Stability: {voice_config.get('stability')}")
        print(f"   Similarity: {voice_config.get('similarityBoost')}")
        print(f"   Style: {voice_config.get('style', 'N/A')}")
        print(f"   Speaker Boost: {voice_config.get('useSpeakerBoost', False)}")
    
    response = requests.post(url, headers=headers, json=payload)
    
    if response.status_code not in [200, 201]:
        error_msg = response.text
        print(f"❌ Failed to create assistant!")
        print(f"   Status: {response.status_code}")
        print(f"   Error: {error_msg}")
        raise Exception(f"Failed to create assistant: {error_msg}")
    
    result = response.json()
    print(f"✅ Assistant created: {result.get('id')}")
    
    return result

# ==========================================
# GET CALL TRANSCRIPT/SUMMARY
# ==========================================
def get_call_transcript(call_id):
    """
    Get call transcript and summary from Vapi API
    Returns conversation transcript and AI-generated summary
    """
    url = f"{VAPI_BASE_URL}/call/{call_id}"
    headers = {
        "Authorization": f"Bearer {VAPI_API_KEY}",
        "Content-Type": "application/json"
    }
    
    try:
        response = requests.get(url, headers=headers)
        
        if response.status_code == 200:
            call_data = response.json()
            
            # Extract transcript
            transcript = call_data.get('transcript', '')
            
            # If transcript is empty, try to get messages
            if not transcript:
                messages = call_data.get('messages', [])
                if messages:
                    # Build transcript from messages
                    transcript_lines = []
                    for msg in messages:
                        role = msg.get('role', 'unknown')
                        content = msg.get('content', '')
                        if content:
                            speaker = "AI" if role == "assistant" else "Student"
                            transcript_lines.append(f"{speaker}: {content}")
                    transcript = "\n".join(transcript_lines)
            
            # Extract summary (if available from Vapi)
            summary = call_data.get('summary', transcript[:200] if transcript else "No transcript available")
            
            # Get analysis data
            analysis = call_data.get('analysis', {})
            
            return {
                'transcript': transcript,
                'summary': summary,
                'analysis': analysis,
                'raw_data': call_data
            }
        else:
            print(f"⚠️ Failed to get transcript: {response.status_code}")
            return {
                'transcript': "Failed to retrieve transcript",
                'summary': "Error fetching call data",
                'analysis': {},
                'raw_data': {}
            }
            
    except Exception as e:
        print(f"❌ Error getting transcript: {e}")
        return {
            'transcript': f"Error: {str(e)}",
            'summary': "Error retrieving call data",
            'analysis': {},
            'raw_data': {}
        }

# ==========================================
# SAVE TO EXCEL
# ==========================================
def save_call_to_excel(phone, language, summary, transcript, duration, cost, status, call_id, filename="call_summaries.xlsx"):
    """
    Append call data to Excel file
    Creates new file if doesn't exist
    """
    try:
        # Create data dictionary
        call_data = {
            'Date & Time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'Phone Number': phone,
            'Language': language,
            'Status': status,
            'Duration': duration,
            'Cost (₹)': cost,
            'Summary': summary,
            'Full Transcript': transcript,
            'Call ID': call_id
        }
        
        # Check if file exists
        if os.path.exists(filename):
            # Append to existing file
            df_existing = pd.read_excel(filename)
            df_new = pd.DataFrame([call_data])
            df_combined = pd.concat([df_existing, df_new], ignore_index=True)
        else:
            # Create new file
            df_combined = pd.DataFrame([call_data])
        
        # Save to Excel
        df_combined.to_excel(filename, index=False, engine='openpyxl')
        
        # Format the Excel file
        wb = load_workbook(filename)
        ws = wb.active
        
        # Set column widths
        ws.column_dimensions['A'].width = 20  # Date & Time
        ws.column_dimensions['B'].width = 15  # Phone
        ws.column_dimensions['C'].width = 12  # Language
        ws.column_dimensions['D'].width = 12  # Status
        ws.column_dimensions['E'].width = 12  # Duration
        ws.column_dimensions['F'].width = 12  # Cost
        ws.column_dimensions['G'].width = 50  # Summary
        ws.column_dimensions['H'].width = 80  # Transcript
        ws.column_dimensions['I'].width = 30  # Call ID
        
        # Style header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Wrap text for summary and transcript columns
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row[6].alignment = Alignment(wrap_text=True, vertical='top')  # Summary
            row[7].alignment = Alignment(wrap_text=True, vertical='top')  # Transcript
        
        wb.save(filename)
        
        print(f"\n✅ Call data saved to {filename}")
        return True
        
    except Exception as e:
        print(f"❌ Error saving to Excel: {e}")
        return False

# ==========================================
# MAKE VAPI CALL
# ==========================================
def make_vapi_call(assistant_id, phone):
    """
    Make outbound call via Vapi
    """
    url = f"{VAPI_BASE_URL}/call/phone"
    
    headers = {
        "Authorization": f"Bearer {VAPI_API_KEY}",
        "Content-Type": "application/json"
    }
    
    payload = {
        "assistantId": assistant_id,
        "phoneNumberId": VAPI_PHONE_NUMBER_ID,
        "customer": {
            "number": phone
        }
    }
    
    print(f"\n📞 Initiating call to {phone}...")
    
    response = requests.post(url, headers=headers, json=payload)
    
    if response.status_code not in [200, 201]:
        error_msg = response.text
        print(f"❌ Call failed!")
        print(f"   Status: {response.status_code}")
        print(f"   Error: {error_msg}")
        raise Exception(f"Failed to make call: {error_msg}")
    
    result = response.json()
    print(f"✅ Call initiated: {result.get('id')}")
    
    return result

# ==========================================
# GET CALL STATUS (ACTUAL)
# ==========================================
def get_call_status(call_id, max_wait=180):
    """
    Poll Vapi API to get actual call status
    Waits up to max_wait seconds for call to complete
    """
    url = f"{VAPI_BASE_URL}/call/{call_id}"
    headers = {
        "Authorization": f"Bearer {VAPI_API_KEY}",
        "Content-Type": "application/json"
    }
    
    import time
    start_time = datetime.now()
    wait_time = 0
    
    print(f"\n⏳ Waiting for call to complete (max {max_wait}s)...")
    
    while wait_time < max_wait:
        try:
            response = requests.get(url, headers=headers)
            
            if response.status_code == 200:
                call_data = response.json()
                status = call_data.get('status', 'unknown')
                
                print(f"   Status: {status} ({wait_time}s elapsed)")
                
                # Check if call ended
                if status in ['ended', 'completed', 'failed', 'busy', 'no-answer']:
                    print(f"\n✅ Call ended with status: {status}")
                    return call_data
                
            # Wait 3 seconds before next check
            time.sleep(3)
            wait_time += 3
            
        except Exception as e:
            print(f"⚠️ Error checking status: {e}")
            time.sleep(3)
            wait_time += 3
    
    print(f"\n⚠️ Timeout: Call still in progress after {max_wait}s")
    return None

# ==========================================
# CALCULATE CALL COST
# ==========================================
def calculate_cost(duration_seconds):
    """
    Calculate call cost based on duration
    Cost breakdown per call: ₹9.13 for ~2 minutes
    """
    if duration_seconds == 0:
        return 4.20  # Minimum VAPI platform fee
    
    duration_minutes = duration_seconds / 60
    
    # Cost per minute calculation
    # ₹9.13 / 2 minutes = ₹4.565 per minute
    cost_per_minute = 4.565
    total_cost = duration_minutes * cost_per_minute
    
    return round(total_cost, 2)

# ==========================================
# MAIN FUNCTION: MAKE CALL WITH LANGUAGE
# ==========================================
def make_call_with_language(language, phone):
    """
    Main function to make call with selected language
    Returns call result with all details
    """
    
    print(f"\n{'='*60}")
    print(f"📞 MAKING CALL")
    print(f"{'='*60}")
    print(f"Language: {language}")
    print(f"Phone: {phone}")
    
    # Get language configuration
    if language not in LANGUAGE_CONFIG:
        raise ValueError(f"Unsupported language: {language}")
    
    config = LANGUAGE_CONFIG[language]
    
    # Create assistant
    voice_params = {
        'stability': config.get('stability', 0.5),
        'similarity_boost': config.get('similarity_boost', 0.75),
        'style': config.get('style'),
        'use_speaker_boost': config.get('use_speaker_boost', False),
        'voice_language': config.get('voice_language')  # For accent control
    }
    
    assistant = create_assistant(
        voice_provider=config['voice_provider'],
        voice_id=config['voice_id'],
        language_code=config['language_code'],
        prompt=config['prompt'],
        language_name=language,
        voice_params=voice_params
    )
    
    # Make the call
    call_result = make_vapi_call(
        assistant_id=assistant['id'],
        phone=phone
    )
    
    call_id = call_result.get('id')
    start_time = datetime.now()
    
    # Wait for call to complete and get actual status
    print(f"\n{'='*60}")
    print(f"⏳ CALL IN PROGRESS - Waiting for completion...")
    print(f"{'='*60}")
    
    final_call_data = get_call_status(call_id, max_wait=180)  # Wait up to 3 minutes
    
    if final_call_data:
        # Get actual duration from API (in seconds)
        duration_seconds = final_call_data.get('duration', 0)
        actual_status = final_call_data.get('status', 'unknown')
        end_time = datetime.now()
        
        # If duration not in response, calculate from timestamps
        if duration_seconds == 0:
            started_at = final_call_data.get('startedAt')
            ended_at = final_call_data.get('endedAt')
            if started_at and ended_at:
                # Parse ISO timestamps and calculate duration
                from dateutil import parser
                start = parser.parse(started_at)
                end = parser.parse(ended_at)
                duration_seconds = int((end - start).total_seconds())
    else:
        # Fallback if timeout
        duration_seconds = 120  # Estimate
        actual_status = 'timeout'
        end_time = datetime.now()
    
    # Calculate actual cost
    cost = calculate_cost(duration_seconds)
    
    print(f"\n{'='*60}")
    print(f"✅ CALL COMPLETED")
    print(f"{'='*60}")
    print(f"Call ID: {call_id}")
    print(f"Status: {actual_status}")
    print(f"Duration: {duration_seconds}s")
    print(f"Cost: ₹{cost}")
    print(f"{'='*60}\n")
    
    # Get call transcript and summary
    print("📝 Fetching call transcript...")
    transcript_data = get_call_transcript(call_id)
    
    # Save to Excel
    print("💾 Saving to Excel...")
    save_call_to_excel(
        phone=phone,
        language=language,
        summary=transcript_data['summary'],
        transcript=transcript_data['transcript'],
        duration=f"{duration_seconds // 60}m {duration_seconds % 60}s",
        cost=cost,
        status=actual_status,
        call_id=call_id
    )
    
    # Return formatted result with ACTUAL data
    return {
        'status': actual_status,  # ACTUAL status from API
        'duration': f"{duration_seconds // 60}m {duration_seconds % 60}s",
        'duration_seconds': duration_seconds,  # ACTUAL duration
        'cost': cost,  # ACTUAL cost based on real duration
        'language': language,
        'call_id': call_id,
        'assistant_id': assistant['id'],
        'recording_url': final_call_data.get('recordingUrl', '') if final_call_data else '',
        'voice_name': config['voice_name'],
        'start_time': start_time.strftime('%Y-%m-%d %H:%M:%S'),
        'end_time': end_time.strftime('%Y-%m-%d %H:%M:%S'),
        'phone': phone,
        'purpose': 'Data Science Feedback Collection',
        'end_reason': final_call_data.get('endedReason', 'unknown') if final_call_data else 'timeout',
        'summary': transcript_data['summary'],
        'transcript': transcript_data['transcript']
    }

# ==========================================
# TEST FUNCTION
# ==========================================
if __name__ == "__main__":
    # Test with sample data
    test_phone = "+919876543210"
    test_language = "English"
    
    print(f"Testing call to {test_phone} in {test_language}...")
    
    try:
        result = make_call_with_language(test_language, test_phone)
        print("\n✅ Call successful!")
        print(f"Call ID: {result['call_id']}")
        print(f"Duration: {result['duration']}")
        print(f"Cost: ₹{result['cost']}")
    except Exception as e:
        print(f"\n❌ Error: {e}")
